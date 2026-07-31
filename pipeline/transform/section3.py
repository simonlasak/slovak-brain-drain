"""
Stage 2 Transform — Section 3: Global diaspora.

Reads from:
- data/raw/oecd/mig_popf_svk.csv (Slovak nationals/born in OECD countries)
- data/raw/oecd/mig_flows_from_svk.csv (migration flows from Slovakia)
- data/raw/un_desa/migrant_stock_bilateral_2020.xlsx (bilateral stock)

Writes to:
- data/processed/section3_diaspora.parquet
"""
from __future__ import annotations

import logging
from pathlib import Path

import polars as pl

log = logging.getLogger(__name__)

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
RAW_OECD = REPO_ROOT / "data" / "raw" / "oecd"
RAW_UN_DESA = REPO_ROOT / "data" / "raw" / "un_desa"

# UN DESA identifies destinations by UN M49 numeric code; OECD uses ISO 3166-1
# alpha-3. `04-spec.md` specifies `destination_iso3`, so ISO3 is the canonical
# key and the UN DESA side is the one that must convert. Emitting both systems
# into one column (introduced in f208339) split 57 real countries across 87 keys,
# double-counting 30 of them and making the frontend join miss the United States
# entirely. Covers every numeric code that appears in the Slovak-origin block.
M49_TO_ISO3 = {
    "036": "AUS", "040": "AUT", "056": "BEL", "068": "BOL", "070": "BIH",
    "076": "BRA", "100": "BGR", "112": "BLR", "124": "CAN", "188": "CRI",
    "191": "HRV", "196": "CYP", "203": "CZE", "208": "DNK", "218": "ECU",
    "233": "EST", "246": "FIN", "250": "FRA", "276": "DEU", "300": "GRC",
    "348": "HUN", "352": "ISL", "372": "IRL", "380": "ITA", "400": "JOR",
    "428": "LVA", "438": "LIE", "440": "LTU", "442": "LUX", "470": "MLT",
    "484": "MEX", "496": "MNG", "499": "MNE", "528": "NLD", "578": "NOR",
    "591": "PAN", "616": "POL", "620": "PRT", "642": "ROU", "643": "RUS",
    "688": "SRB", "705": "SVN", "710": "ZAF", "724": "ESP", "752": "SWE",
    "756": "CHE", "792": "TUR", "807": "MKD", "818": "EGY", "826": "GBR",
    "862": "VEN",
}

# OECD SDMX observation status. Only "A" (normal) is an actual observation; the
# rest indicate the value was modelled, provisional, or otherwise not a direct
# measurement. `03-methodology.md` display honesty principle 2 requires actual
# and interpolated points to be distinguished visually, which needs this flag
# carried rather than hardcoded. Anything not in this map yields null, not false:
# "we do not know" is not the same claim as "this is an observation".
OBS_STATUS_INTERPOLATED = {
    "A": False,   # Normal value
    "E": True,    # Estimated
    "P": False,   # Provisional, but still an observation
    "I": True,    # Imputed
    "M": None,    # Missing
    "L": True,    # Missing, data exist but were not collected
}
OUT_PATH = REPO_ROOT / "data" / "processed" / "section3_diaspora.parquet"


def transform_oecd_popf() -> pl.DataFrame:
    """OECD population of foreign/foreign-born — Slovaks in OECD countries."""
    fpath = RAW_OECD / "mig_popf_svk_abroad.csv"
    if not fpath.exists():
        fpath = RAW_OECD / "mig_popf_svk.csv"
    df = pl.read_csv(fpath, infer_schema_length=10000)

    rows = []
    for row in df.iter_rows(named=True):
        year = row.get("TIME_PERIOD")
        if year is None:
            continue
        year = int(year)
        if year < 1990:
            continue

        destination = row.get("REF_AREA", "")
        if not destination or destination == "SVK":
            continue

        value = row.get("OBS_VALUE")
        if value is None or float(value) == 0:
            continue

        sex_raw = row.get("SEX", "_T")
        sex = {"_T": "all", "M": "M", "F": "F"}.get(str(sex_raw), "all")

        measure = str(row.get("MEASURE", "") or "").strip()
        obs_status = str(row.get("OBS_STATUS", "") or "").strip()

        rows.append({
            "year": year,
            "slovak_def": "born",
            "destination_iso3": str(destination),
            "sex": sex,
            "age_bracket": "all",
            "education": "all",
            "metric": "stock",
            "value": float(value),
            "is_interpolated": OBS_STATUS_INTERPOLATED.get(obs_status),
            # Derived per-observation from the file's own OBS_STATUS column.
            "flag_basis": "observation",
            "source": "oecd_mig_popf",
            "measure_code": measure or None,
            "obs_status": obs_status or None,
            # UN DESA-only provenance; null for OECD rows.
            "data_type": None,
            "data_type_note": None,
        })

    return pl.DataFrame(rows)


# OECD DSD_MIG MEASURE codes present in mig_flows_svk_abroad.csv. The file
# stacks five different measures in one table, distinguished ONLY by this
# column: every other dimension (year, citizenship, sex, age, education, unit)
# is identical across them. Reading OBS_VALUE without filtering on MEASURE
# therefore produced several contradictory rows per country-year and silently
# mixed stocks into the flow series.
#
# Identified by cross-validation against independent sources:
#   B11 inflow of Slovak citizens   CZE ~5.8-7.2k/yr, matches the "6,000 to
#                                   7,000 per year" figure Section 2 takes
#                                   from OECD commentary
#   B12 outflow / departures        smaller counter-flow
#   B13 asylum applications         effectively zero for Slovaks
#   B14 stock of foreign population (in the popf file, not this one)
#   B15 stock of Slovak population  CZE series equals CSU CIZ003T003 exactly
#                                   but shifted one year (start- vs end-of-year
#                                   convention), so it is NOT a flow
#   B16 naturalisations             low hundreds
#
# Only B11 is a genuine inflow of Slovaks, so only B11 becomes metric='inflow'.
# B12 and B16 are kept under their own metric names because they are real and
# may be wanted later; B13 and B15 are dropped (B13 is noise, B15 duplicates
# Section 2's residence series on a different year convention).
OECD_FLOW_MEASURES = {
    "B11": "inflow",
    "B12": "outflow",
    "B16": "naturalisations",
}

# Greece reports B15 for 2022 as 780,000, which is three orders of magnitude
# above every neighbouring observation and above the entire Slovak diaspora.
# It is a unit error (thousands reported as persons) in the upstream file. B15
# is dropped anyway, but guard the threshold so no future measure smuggles it
# back in.
IMPLAUSIBLE_VALUE = 500_000


def transform_oecd_flows() -> pl.DataFrame:
    """OECD migration flows — Slovaks arriving in and leaving OECD countries.

    Splits the stacked MEASURE codes into distinct metrics instead of
    collapsing them into a single undifferentiated 'inflow'.
    """
    fpath = RAW_OECD / "mig_flows_svk_abroad.csv"
    if not fpath.exists():
        fpath = RAW_OECD / "mig_flows_from_svk.csv"
    df = pl.read_csv(fpath, infer_schema_length=10000)

    if "MEASURE" not in df.columns:
        raise ValueError(
            f"{fpath.name} has no MEASURE column; the OECD schema has changed "
            "and the measures can no longer be separated. Refusing to emit a "
            "silently mixed series."
        )

    rows = []
    dropped: dict[str, int] = {}
    for row in df.iter_rows(named=True):
        year = row.get("TIME_PERIOD")
        if year is None or str(year).strip() == "":
            continue
        year = int(year)
        if year < 1990:
            continue

        destination = row.get("REF_AREA", "")
        if not destination or destination == "SVK":
            continue

        measure = str(row.get("MEASURE", "")).strip()
        metric = OECD_FLOW_MEASURES.get(measure)
        if metric is None:
            dropped[measure] = dropped.get(measure, 0) + 1
            continue

        value = row.get("OBS_VALUE")
        if value is None or str(value).strip() == "" or float(value) == 0:
            continue
        value = float(value)
        if value > IMPLAUSIBLE_VALUE:
            log.warning(
                "transform.section3.implausible_value measure=%s dest=%s year=%s value=%s (dropped)",
                measure, destination, year, value,
            )
            continue

        sex_raw = row.get("SEX", "_T")
        sex = {"_T": "all", "M": "M", "F": "F"}.get(str(sex_raw), "all")
        obs_status = str(row.get("OBS_STATUS", "") or "").strip()

        rows.append({
            "year": year,
            "slovak_def": "citizen",
            "destination_iso3": str(destination),
            "sex": sex,
            "age_bracket": "all",
            "education": "all",
            "metric": metric,
            "value": value,
            "is_interpolated": OBS_STATUS_INTERPOLATED.get(obs_status),
            # Derived per-observation from the file's own OBS_STATUS column.
            "flag_basis": "observation",
            "source": f"oecd_mig_flows_{measure}",
            "measure_code": measure,
            "obs_status": obs_status or None,
            "data_type": None,
            "data_type_note": None,
        })

    if dropped:
        log.info("transform.section3.flows_measures_dropped %s", dropped)

    return pl.DataFrame(rows)


# UN DESA location code for Slovakia, used to select the origin row.
# UN DESA "Type of data of destination" codes, documented on the workbook's own
# Notes sheet. Preserved as a field because the codes are not interchangeable:
# Czechia is the single C row among the 51 Slovak-origin destinations, so the
# site's largest figure is a CITIZEN count sitting beside 50 birth-based ones.
# Flattening this column away is what hid that.
DESA_DATA_TYPES = {
    "B": "Derived from data on the foreign-born population",
    "C": "Derived from data on foreign citizens",
    "R": "Includes refugees, asylum seekers or Venezuelans displaced abroad",
    "I": "Imputed from a regional or country model",
}

SVK_LOCATION_CODE = 703
# Aggregate rows (WORLD, regions, development groups) share the 900+ code space
# and must not be mistaken for destination countries.
UN_AGGREGATE_MIN = 900


def transform_un_desa() -> pl.DataFrame:
    """UN DESA bilateral migrant stock — Slovak-born by destination country.

    Table 1 of the 2020 revision is a destination-by-origin matrix: one row per
    destination, one column per origin, with a block of columns per reference
    year. We take the Slovakia origin column for every destination row.

    This is the only source with broad country coverage (51 destinations in
    2020 versus roughly 22 from OECD), which is why Section 3's map uses a UN
    DESA snapshot year rather than an annual series.
    """
    fpath = RAW_UN_DESA / "migrant_stock_bilateral_2020.xlsx"
    if not fpath.exists():
        log.warning("transform.section3.un_desa_missing path=%s", fpath)
        return pl.DataFrame()

    import openpyxl

    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    ws = wb["Table 1"]

    rows: list[dict] = []
    header: list | None = None
    # (sex, year) -> column index. Table 1 repeats the whole year block three
    # times: both sexes, then males, then females, labelled only in a banner row
    # above the year header. Keying by year alone would let the male and female
    # blocks overwrite the combined one and silently halve every figure.
    sex_year_cols: dict[tuple[str, int], int] = {}
    sex_blocks: list[tuple[int, str]] = []
    unmapped_codes: set[str] = set()

    for raw in ws.iter_rows(values_only=True):
        if header is None:
            # Banner row naming each block; capture where each sex starts.
            for idx, cell in enumerate(raw or ()):
                if cell is None:
                    continue
                text = str(cell).lower()
                if "international migrant stock" not in text:
                    continue
                if "both sexes" in text:
                    sex_blocks.append((idx, "all"))
                elif "male" in text and "female" not in text:
                    sex_blocks.append((idx, "M"))
                elif "female" in text:
                    sex_blocks.append((idx, "F"))

            # The sheet carries several rows of citation preamble; the real
            # header is the row whose first cell is "Index".
            if raw and str(raw[0]).strip() == "Index":
                header = list(raw)
                if not sex_blocks:
                    raise ValueError(
                        "UN DESA Table 1: could not locate the sex block banners, "
                        "so year columns cannot be attributed to a sex."
                    )
                sex_blocks.sort()
                for idx, cell in enumerate(header):
                    text = str(cell).strip() if cell is not None else ""
                    if not (text.isdigit() and 1990 <= int(text) <= 2100):
                        continue
                    # Attribute this year column to the last block that starts
                    # at or before it.
                    sex = None
                    for start, label in sex_blocks:
                        if idx >= start:
                            sex = label
                    if sex is not None:
                        sex_year_cols[(sex, int(text))] = idx
                if not sex_year_cols:
                    raise ValueError("UN DESA Table 1: no year columns found")
            continue

        # Columns are positional: 3 = destination location code,
        # 6 = origin location code (0-indexed).
        dest_code = raw[3]
        origin_code = raw[6]
        if dest_code is None or origin_code is None:
            continue
        try:
            dest_code = int(dest_code)
            origin_code = int(origin_code)
        except (TypeError, ValueError):
            continue

        if origin_code != SVK_LOCATION_CODE:
            continue
        # Skip WORLD / regional / development-group aggregate destinations.
        if dest_code >= UN_AGGREGATE_MIN:
            continue

        # Canonicalise to ISO3 so the column holds one code system. A numeric
        # code with no mapping is skipped rather than emitted raw, which would
        # silently reintroduce the two-key split.
        m49 = str(dest_code).zfill(3)
        iso3 = M49_TO_ISO3.get(m49)
        if iso3 is None:
            unmapped_codes.add(m49)
            continue

        # Column 4 is "Type of data of destination": one or more of B/C/R/I.
        raw_type = str(raw[4] or "").strip()
        type_codes = [t for t in raw_type.split() if t in DESA_DATA_TYPES]
        data_type = " ".join(type_codes) or None
        data_type_note = (
            "; ".join(DESA_DATA_TYPES[t] for t in type_codes) or None
        )

        for (sex, year), col in sex_year_cols.items():
            value = raw[col] if col < len(raw) else None
            if value is None:
                continue
            try:
                value = float(value)
            except (TypeError, ValueError):
                continue
            if value <= 0:
                continue

            rows.append({
                "year": year,
                "slovak_def": "born",
                "destination_iso3": iso3,
                "sex": sex,
                "age_bracket": "all",
                "education": "all",
                "metric": "stock",
                "value": value,
                # NOT DERIVED FROM THE FILE. This is a constant, set because the
                # 2020 revision Methodology Report section 6 states that every
                # reference year is produced by interpolation or extrapolation.
                # The bilateral matrix itself carries no per-observation status
                # column, so nothing here can confirm or contradict it: the flag
                # restates a document.
                #
                # Consequence for copy: a claim that "the parquet flags all seven
                # reference years as interpolated" is circular. Cite the
                # methodology report directly instead. Kept as True because it is
                # the right value and downstream display logic depends on it, but
                # it is provenance-by-assertion, not evidence.
                "is_interpolated": True,
                # NOT per-observation. See the comment above: this restates the
                # methodology report, so the flag is a source-level assertion.
                "flag_basis": "source_document",
                "source": "un_desa_bilateral_2020",
                "measure_code": None,
                "obs_status": None,
                "data_type": data_type,
                "data_type_note": data_type_note,
            })

    wb.close()
    if unmapped_codes:
        log.warning(
            "transform.section3.un_desa_unmapped_m49 %s (rows dropped; add to M49_TO_ISO3)",
            sorted(unmapped_codes),
        )
    return pl.DataFrame(rows)


# Eurostat migr_pop1ctz: destination-reported stock of Slovak CITIZENS. This is
# the second of the three definitions 01-research-architecture.md:84 asks §3 to
# compare, and it was previously absent from this parquet: slovak_def='citizen'
# existed only for FLOWS (inflow/outflow/naturalisations), so a definitional
# comparison of STOCKS could not be built from one table.
#
# Slovakia reports its own resident Slovak citizens in this dataset. That row is
# excluded: it counts people who never left.
EUROSTAT_SLOVAK_CITIZEN = "SK"
RAW_EUROSTAT = REPO_ROOT / "data" / "raw" / "eurostat"

# Eurostat uses 2-letter geo codes; this column is destination_iso3. Emitting
# Eurostat codes raw would put two code systems in one column, which is exactly
# the defect the M49_TO_ISO3 comment above records (57 real countries split
# across 87 keys). Unmapped codes are logged and dropped, never passed through.
#
# EL is Greece and UK is the United Kingdom in Eurostat's coding; both differ
# from ISO 3166-1. XK is Kosovo, which has no ISO3 assignment.
EUROSTAT_GEO_TO_ISO3 = {
    "AT": "AUT", "BE": "BEL", "BG": "BGR", "CH": "CHE", "CY": "CYP",
    "CZ": "CZE", "DE": "DEU", "DK": "DNK", "EE": "EST", "EL": "GRC",
    "ES": "ESP", "FI": "FIN", "FR": "FRA", "HR": "HRV", "HU": "HUN",
    "IE": "IRL", "IS": "ISL", "IT": "ITA", "LI": "LIE", "LT": "LTU",
    "LU": "LUX", "LV": "LVA", "ME": "MNE", "MK": "MKD", "MT": "MLT",
    "NL": "NLD", "NO": "NOR", "PL": "POL", "PT": "PRT", "RO": "ROU",
    "RS": "SRB", "SE": "SWE", "SI": "SVN", "TR": "TUR", "UK": "GBR",
    "AL": "ALB", "BA": "BIH", "BY": "BLR", "MD": "MDA", "UA": "UKR",
    "AM": "ARM", "AZ": "AZE", "GE": "GEO", "RU": "RUS",
}


def transform_eurostat_citizen_stock() -> pl.DataFrame:
    """Eurostat migr_pop1ctz — Slovak citizens resident in each reporting country.

    Parsed here rather than imported from pipeline.analysis.mirror_comparison so
    the transform layer has no dependency on the analysis layer, but the filter
    is deliberately identical: citizen=SK, age=TOTAL, sex=T, unit=NR.
    """
    import gzip

    path = RAW_EUROSTAT / "migr_pop1ctz.tsv.gz"
    if not path.exists():
        log.warning("transform.section3.citizen_stock: migr_pop1ctz absent, skipping")
        return pl.DataFrame()

    rows = []
    seen: set[tuple[str, int]] = set()
    unmapped_geo: set[str] = set()
    with gzip.open(path, "rt", encoding="utf-8") as fh:
        header = fh.readline().rstrip("\n").split("\t")
        dim_names = header[0].split("\\")[0].split(",")
        years = [int(h.strip()) for h in header[1:] if h.strip()]

        for line in fh:
            if not line.strip():
                continue
            cells = line.rstrip("\n").split("\t")
            dims = cells[0].split(",")
            if len(dims) != len(dim_names):
                continue
            d = dict(zip(dim_names, dims))
            if not (d.get("citizen") == EUROSTAT_SLOVAK_CITIZEN
                    and d.get("age") == "TOTAL"
                    and d.get("sex") == "T"
                    and d.get("unit") == "NR"):
                continue
            geo_raw = d.get("geo", "")
            if not geo_raw or geo_raw == EUROSTAT_SLOVAK_CITIZEN:
                continue
            geo = EUROSTAT_GEO_TO_ISO3.get(geo_raw)
            if geo is None:
                unmapped_geo.add(geo_raw)
                continue

            for year, raw in zip(years, cells[1:]):
                token = raw.strip()
                if not token or token.startswith(":"):
                    continue
                # Eurostat appends observation flags to the number ("4 b", "12 p").
                # Strip before casting or every flagged value becomes null.
                number = token.split(" ")[0].replace(",", "")
                try:
                    value = float(number)
                except ValueError:
                    continue
                if value <= 0:
                    continue
                key = (geo, year)
                if key in seen:
                    raise ValueError(
                        f"section3: migr_pop1ctz yields two rows for {geo} {year} "
                        "after filtering to Slovak citizens, all ages, both sexes, "
                        "persons. A dimension is unconstrained."
                    )
                seen.add(key)

                # Eurostat flags: 'p' provisional, 'e' estimated, 'b' break in
                # series, 'd' definition differs. Only 'e' asserts the value is
                # not a direct observation.
                flags = token.split(" ")[1] if " " in token else ""
                rows.append({
                    "year": year,
                    "slovak_def": "citizen",
                    "destination_iso3": geo,
                    "sex": "all",
                    "age_bracket": "all",
                    "education": "all",
                    "metric": "stock",
                    "value": value,
                    "is_interpolated": True if "e" in flags else False,
                    # Per-observation, from the file's own flag characters.
                    "flag_basis": "observation",
                    "source": "eurostat_migr_pop1ctz",
                    "measure_code": None,
                    "obs_status": flags or None,
                    # Eurostat reports by citizenship by construction, which is
                    # the definitional counterpart to UN DESA's type-B/C split.
                    "data_type": "C",
                    "data_type_note": "Reported by current citizenship (Eurostat migr_pop1ctz)",
                })

    if not rows:
        raise ValueError(
            "section3: migr_pop1ctz parsed but yielded no Slovak-citizen rows. "
            "The citizenship code or a dimension name changed."
        )
    if unmapped_geo:
        log.warning(
            "transform.section3.citizen_stock_unmapped_geo %s (dropped, not "
            "passed through: mixing code systems in destination_iso3 is what "
            "split 57 countries across 87 keys)", sorted(unmapped_geo))
    log.info("transform.section3.citizen_stock countries=%d",
             len({r["destination_iso3"] for r in rows}))
    return pl.DataFrame(rows)


def run() -> pl.DataFrame:
    log.info("transform.section3.start")

    df_popf = transform_oecd_popf()
    log.info("transform.section3.popf rows=%d", len(df_popf))

    df_flows = transform_oecd_flows()
    log.info("transform.section3.flows rows=%d", len(df_flows))

    df_desa = transform_un_desa()
    log.info("transform.section3.un_desa rows=%d", len(df_desa))

    df_citizen = transform_eurostat_citizen_stock()
    log.info("transform.section3.citizen_stock rows=%d", len(df_citizen))

    frames = [f for f in [df_popf, df_flows, df_desa, df_citizen] if len(f) > 0]
    df = pl.concat(frames, how="vertical_relaxed")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.write_parquet(OUT_PATH)
    log.info("transform.section3.done rows=%d path=%s", len(df), OUT_PATH)
    return df


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO, stream=sys.stderr)
    df = run()
    print(f"Section 3: {len(df):,} rows written to {OUT_PATH}")
    print(f"Metrics: {df['metric'].unique().sort().to_list()}")
    print(f"Years: {df['year'].min()}–{df['year'].max()}")
    print(f"Destinations: {df['destination_iso3'].n_unique()}")
    print(f"Top destinations (stock):")
    top = (df.filter(pl.col('metric') == 'stock')
           .group_by('destination_iso3')
           .agg(pl.col('value').max())
           .sort('value', descending=True)
           .head(10))
    print(top)
